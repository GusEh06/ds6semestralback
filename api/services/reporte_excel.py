# services/reporte_excel.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, time, date
from django.db.models import Count, Avg, Prefetch
from django.db import connection
from ..models import (
    Visitante, Sendero, RegistroVisita, 
    Encuesta, Usuario, Comentario
)
import gc


def generar_reporte_completo():
    """Genera UN SOLO reporte completo con todas las hojas necesarias - OPTIMIZADO."""
    
    try:
        # Configurar límites para evitar problemas de memoria
        MAX_ROWS_PER_SHEET = 10000
        
        wb = openpyxl.Workbook()
        
        # Estilos reutilizables (más simples para reducir memoria)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        def aplicar_estilos_header_simple(ws, headers):
            """Aplica estilos simples a los headers."""
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.value = header
                # Ajustar ancho de columna de forma más eficiente
                ws.column_dimensions[get_column_letter(col_num)].width = min(max(len(str(header)) + 2, 12), 30)

        def formatear_fecha(fecha_obj):
            """Formatea fecha de manera segura y eficiente."""
            if not fecha_obj:
                return 'N/A'
            try:
                if isinstance(fecha_obj, datetime):
                    return fecha_obj.strftime("%d/%m/%Y")
                elif isinstance(fecha_obj, date):
                    return fecha_obj.strftime("%d/%m/%Y")
                else:
                    return str(fecha_obj)
            except:
                return 'N/A'

        def formatear_fecha_hora(fecha_obj):
            """Formatea fecha con hora de manera segura y eficiente."""
            if not fecha_obj:
                return 'N/A'
            try:
                if isinstance(fecha_obj, datetime):
                    return fecha_obj.strftime("%d/%m/%Y %H:%M")
                elif isinstance(fecha_obj, date):
                    return fecha_obj.strftime("%d/%m/%Y")
                else:
                    return str(fecha_obj)
            except:
                return 'N/A'

        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # ===== HOJA 1: RESUMEN EJECUTIVO =====
        ws_resumen = wb.create_sheet(title="1. Resumen")
        
        # Obtener estadísticas con una sola consulta optimizada
        try:
            total_visitantes = Visitante.objects.count()
            total_visitas = RegistroVisita.objects.count()
            total_senderos = Sendero.objects.count()
            total_encuestas = Encuesta.objects.count()
            total_usuarios = Usuario.objects.count()
            total_comentarios = Comentario.objects.count()
        except Exception as e:
            # Si falla la consulta, usar valores por defecto
            total_visitantes = total_visitas = total_senderos = 0
            total_encuestas = total_usuarios = total_comentarios = 0

        # Título simple
        ws_resumen['A1'] = "REPORTE CENTRO DE VISITANTES"
        ws_resumen['A1'].font = Font(size=14, bold=True)
        
        ws_resumen['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Métricas principales (más compacto)
        metricas = [
            ("Total Visitantes", total_visitantes),
            ("Total Visitas", total_visitas),
            ("Total Senderos", total_senderos),
            ("Total Encuestas", total_encuestas),
            ("Total Usuarios", total_usuarios),
            ("Total Comentarios", total_comentarios),
        ]
        
        row = 5
        for metrica, valor in metricas:
            ws_resumen[f'A{row}'] = metrica
            ws_resumen[f'B{row}'] = valor
            ws_resumen[f'A{row}'].font = Font(bold=True)
            row += 1

        # ===== HOJA 2: VISITANTES (LIMITADO) =====
        ws_visitantes = wb.create_sheet(title="2. Visitantes")
        
        headers_visitantes = ['ID', 'Nombre', 'Documento', 'Nacionalidad', 'Tipo', 'Género']
        ws_visitantes.append(headers_visitantes)
        aplicar_estilos_header_simple(ws_visitantes, headers_visitantes)
        
        # Usar select_related y limit para optimizar
        visitantes = Visitante.objects.all()[:MAX_ROWS_PER_SHEET]
        
        for visitante in visitantes:
            try:
                ws_visitantes.append([
                    visitante.id,
                    (visitante.nombre_visitante or 'N/A')[:50],  # Truncar textos largos
                    (visitante.cedula_pasaporte or 'N/A')[:20],
                    (visitante.nacionalidad or 'N/A')[:30],
                    (visitante.adulto_nino or 'N/A')[:10],
                    (visitante.genero or 'N/A')[:10],
                ])
            except Exception:
                continue  # Saltar registros problemáticos

        # Limpiar memoria
        del visitantes
        gc.collect()

        # ===== HOJA 3: VISITAS RECIENTES (LIMITADO) =====
        ws_visitas = wb.create_sheet(title="3. Visitas")
        
        headers_visitas = ['ID', 'Fecha', 'Visitante', 'Sendero', 'Razón']
        ws_visitas.append(headers_visitas)
        aplicar_estilos_header_simple(ws_visitas, headers_visitas)
        
        # Obtener solo las visitas más recientes para reducir carga
        visitas = RegistroVisita.objects.select_related('visitante').order_by('-fecha_visita')[:MAX_ROWS_PER_SHEET]
        
        for visita in visitas:
            try:
                fecha_formateada = formatear_fecha(visita.fecha_visita)
                visitante_nombre = 'N/A'
                if visita.visitante:
                    visitante_nombre = (visita.visitante.nombre_visitante or 'N/A')[:30]

                ws_visitas.append([
                    visita.id,
                    fecha_formateada,
                    visitante_nombre,
                    (visita.sendero_visitado or 'N/A')[:40],
                    (visita.razon_visita or 'N/A')[:60],
                ])
            except Exception:
                continue

        del visitas
        gc.collect()

        # ===== HOJA 4: SENDEROS =====
        ws_senderos = wb.create_sheet(title="4. Senderos")
        
        headers_senderos = ['ID', 'Nombre', 'Distancia', 'Dificultad', 'Visitas']
        ws_senderos.append(headers_senderos)
        aplicar_estilos_header_simple(ws_senderos, headers_senderos)
        
        senderos = Sendero.objects.all()
        
        for sendero in senderos:
            try:
                # Contar visitas de forma más eficiente
                total_visitas = RegistroVisita.objects.filter(
                    sendero_visitado__icontains=sendero.nombre_sendero
                ).count()
                
                ws_senderos.append([
                    sendero.id,
                    sendero.nombre_sendero[:50],
                    float(sendero.distancia) if sendero.distancia else 0,
                    sendero.dificultad[:20],
                    total_visitas
                ])
            except Exception:
                continue

        del senderos
        gc.collect()

        # ===== HOJA 5: ESTADÍSTICAS POR PAÍS =====
        ws_paises = wb.create_sheet(title="5. Por País")
        
        headers_paises = ['País', 'Visitantes', 'Porcentaje']
        ws_paises.append(headers_paises)
        aplicar_estilos_header_simple(ws_paises, headers_paises)
        
        try:
            # Consulta más eficiente
            stats_pais = Visitante.objects.values('nacionalidad').annotate(
                total=Count('id')
            ).order_by('-total')[:50]  # Limitar a top 50 países
            
            total_general = Visitante.objects.count()
            
            for stat in stats_pais:
                if total_general > 0:
                    porcentaje = (stat['total'] / total_general * 100)
                    ws_paises.append([
                        (stat['nacionalidad'] or 'No especificado')[:30],
                        stat['total'],
                        f"{porcentaje:.1f}%"
                    ])
        except Exception:
            ws_paises.append(['Error al cargar datos', '', ''])

        # ===== HOJA 6: USUARIOS =====
        ws_usuarios = wb.create_sheet(title="6. Usuarios")
        
        headers_usuarios = ['ID', 'Email', 'Nombre', 'Rol']
        ws_usuarios.append(headers_usuarios)
        aplicar_estilos_header_simple(ws_usuarios, headers_usuarios)
        
        usuarios = Usuario.objects.all()[:MAX_ROWS_PER_SHEET]
        
        for usuario in usuarios:
            try:
                ws_usuarios.append([
                    usuario.id,
                    (usuario.email or 'N/A')[:50],
                    f"{usuario.nombre or ''} {usuario.apellido or ''}".strip()[:40] or 'N/A',
                    usuario.get_rol_display()[:20] if hasattr(usuario, 'get_rol_display') else 'N/A'
                ])
            except Exception:
                continue

        del usuarios
        gc.collect()

        # Generar archivo de forma más eficiente
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Limpiar memoria
        wb.close()
        gc.collect()
        
        return output
        
    except Exception as e:
        # Log del error para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generando reporte Excel: {str(e)}", exc_info=True)
        
        # Crear un reporte simple de error
        wb_error = openpyxl.Workbook()
        ws_error = wb_error.active
        ws_error.title = "Error"
        ws_error['A1'] = f"Error: {str(e)[:100]}"
        ws_error['A2'] = "Contacte al administrador"
        
        output_error = BytesIO()
        wb_error.save(output_error)
        output_error.seek(0)
        wb_error.close()
        
        return output_error

    finally:
        # Cerrar conexiones de DB si es necesario
        connection.close()
