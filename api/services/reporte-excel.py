# services/reporte_excel.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
from django.db.models import Count, Avg
from ..models import (
    Visitante, Sendero, RegistroVisita, 
    Encuesta, Usuario, Comentario
)


def generar_reporte_completo():
    """Genera UN SOLO reporte completo con todas las hojas necesarias."""
    
    wb = openpyxl.Workbook()
    
    # Estilos reutilizables
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    def aplicar_estilos_header(ws, headers):
        """Aplica estilos a los headers."""
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = max(len(str(header)) + 3, 15)

    def aplicar_bordes(ws, max_row, max_col):
        """Aplica bordes a todas las celdas."""
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = border

    # Eliminar hoja por defecto
    wb.remove(wb.active)

    # ===== HOJA 1: RESUMEN EJECUTIVO =====
    ws_resumen = wb.create_sheet(title="1. Resumen Ejecutivo")
    
    # Título
    ws_resumen.merge_cells('A1:D1')
    titulo = ws_resumen['A1']
    titulo.value = "REPORTE COMPLETO - CENTRO DE VISITANTES"
    titulo.font = Font(size=16, bold=True, color="FFFFFF")
    titulo.alignment = center_alignment
    titulo.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    ws_resumen['A3'] = f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumen['A3'].font = Font(bold=True)
    
    # Métricas principales
    metricas = [
        ("📊 ESTADÍSTICAS GENERALES", ""),
        ("Total de Visitantes Registrados", Visitante.objects.count()),
        ("Total de Visitas Realizadas", RegistroVisita.objects.count()),
        ("Total de Senderos Disponibles", Sendero.objects.count()),
        ("Total de Encuestas Completadas", Encuesta.objects.count()),
        ("Total de Usuarios del Sistema", Usuario.objects.count()),
        ("Total de Comentarios", Comentario.objects.count()),
    ]
    
    row = 5
    for metrica, valor in metricas:
        ws_resumen[f'A{row}'] = metrica
        if valor != "":
            ws_resumen[f'B{row}'] = valor
        if metrica.startswith("📊"):
            ws_resumen[f'A{row}'].font = Font(bold=True, size=14)
        else:
            ws_resumen[f'A{row}'].font = Font(bold=True)
        row += 1

    # ===== HOJA 2: VISITANTES =====
    ws_visitantes = wb.create_sheet(title="2. Visitantes")
    
    headers_visitantes = [
        'ID', 'Nombre Completo', 'Cédula/Pasaporte', 'Nacionalidad', 
        'Tipo (Adulto/Niño)', 'Género', 'Teléfono', 'Total de Visitas'
    ]
    ws_visitantes.append(headers_visitantes)
    aplicar_estilos_header(ws_visitantes, headers_visitantes)
    
    for visitante in Visitante.objects.all():
        total_visitas = RegistroVisita.objects.filter(visitante=visitante).count()
        ws_visitantes.append([
            visitante.id,
            visitante.nombre_visitante or 'N/A',
            visitante.cedula_pasaporte or 'N/A',
            visitante.nacionalidad or 'N/A',
            visitante.adulto_nino or 'N/A',
            visitante.genero or 'N/A',
            visitante.telefono or 'N/A',
            total_visitas
        ])
    
    aplicar_bordes(ws_visitantes, ws_visitantes.max_row, len(headers_visitantes))

    # ===== HOJA 3: REGISTRO DE VISITAS =====
    ws_visitas = wb.create_sheet(title="3. Registro de Visitas")
    
    headers_visitas = [
        'ID Visita', 'Fecha', 'Hora Entrada', 'Nombre Visitante', 
        'Nacionalidad', 'Sendero Visitado', 'Razón de la Visita', 'Tipo Visitante'
    ]
    ws_visitas.append(headers_visitas)
    aplicar_estilos_header(ws_visitas, headers_visitas)
    
    for visita in RegistroVisita.objects.select_related('visitante').order_by('-fecha_visita'):
        ws_visitas.append([
            visita.id,
            visita.fecha_visita.strftime("%d/%m/%Y"),
            visita.hora_entrada.strftime("%H:%M") if visita.hora_entrada else 'N/A',
            visita.visitante.nombre_visitante or 'N/A',
            visita.visitante.nacionalidad or 'N/A',
            visita.sendero_visitado or 'N/A',
            (visita.razon_visita[:50] + '...') if len(visita.razon_visita or '') > 50 else (visita.razon_visita or 'N/A'),
            visita.visitante.adulto_nino or 'N/A'
        ])
    
    aplicar_bordes(ws_visitas, ws_visitas.max_row, len(headers_visitas))

    # ===== HOJA 4: SENDEROS =====
    ws_senderos = wb.create_sheet(title="4. Senderos")
    
    headers_senderos = [
        'ID', 'Nombre del Sendero', 'Distancia (km)', 'Nivel de Dificultad', 
        'Total de Visitas', 'Promedio de Valoración', 'Total de Comentarios'
    ]
    ws_senderos.append(headers_senderos)
    aplicar_estilos_header(ws_senderos, headers_senderos)
    
    for sendero in Sendero.objects.all():
        # Contar visitas que mencionen este sendero
        total_visitas = RegistroVisita.objects.filter(
            sendero_visitado__icontains=sendero.nombre_sendero
        ).count()
        
        # Estadísticas de comentarios
        comentarios = Comentario.objects.filter(sendero=sendero)
        total_comentarios = comentarios.count()
        promedio_valoracion = comentarios.aggregate(
            promedio=Avg('valoracion')
        )['promedio'] or 0
        
        ws_senderos.append([
            sendero.id,
            sendero.nombre_sendero,
            float(sendero.distancia),
            sendero.dificultad,
            total_visitas,
            round(promedio_valoracion, 1) if promedio_valoracion else 'N/A',
            total_comentarios
        ])
    
    aplicar_bordes(ws_senderos, ws_senderos.max_row, len(headers_senderos))

    # ===== HOJA 5: ENCUESTAS =====
    ws_encuestas = wb.create_sheet(title="5. Encuestas")
    
    headers_encuestas = [
        'ID', 'Fecha de la Encuesta', 'Nombre del Visitante', 
        'Sendero Visitado', 'Respuestas del Formulario'
    ]
    ws_encuestas.append(headers_encuestas)
    aplicar_estilos_header(ws_encuestas, headers_encuestas)
    
    for encuesta in Encuesta.objects.select_related('visita__visitante').order_by('-fecha_visita'):
        # Convertir JSON a texto legible
        formulario_texto = str(encuesta.formulario)
        if len(formulario_texto) > 100:
            formulario_texto = formulario_texto[:100] + '...'
        
        ws_encuestas.append([
            encuesta.id,
            encuesta.fecha_visita.strftime("%d/%m/%Y %H:%M"),
            encuesta.visita.visitante.nombre_visitante or 'N/A',
            encuesta.visita.sendero_visitado or 'N/A',
            formulario_texto
        ])
    
    aplicar_bordes(ws_encuestas, ws_encuestas.max_row, len(headers_encuestas))

    # ===== HOJA 6: ESTADÍSTICAS POR PAÍS =====
    ws_paises = wb.create_sheet(title="6. Estadísticas por País")
    
    headers_paises = ['País/Nacionalidad', 'Total Visitantes', 'Total Visitas', 'Porcentaje del Total']
    ws_paises.append(headers_paises)
    aplicar_estilos_header(ws_paises, headers_paises)
    
    # Estadísticas por nacionalidad
    stats_pais = Visitante.objects.values('nacionalidad').annotate(
        total_visitantes=Count('id')
    ).order_by('-total_visitantes')
    
    total_general = Visitante.objects.count()
    
    for stat in stats_pais:
        # Contar visitas por país
        visitantes_pais = Visitante.objects.filter(nacionalidad=stat['nacionalidad'])
        total_visitas = RegistroVisita.objects.filter(visitante__in=visitantes_pais).count()
        
        porcentaje = (stat['total_visitantes'] / total_general * 100) if total_general > 0 else 0
        
        ws_paises.append([
            stat['nacionalidad'] or 'No especificado',
            stat['total_visitantes'],
            total_visitas,
            f"{porcentaje:.1f}%"
        ])
    
    aplicar_bordes(ws_paises, ws_paises.max_row, len(headers_paises))

    # ===== HOJA 7: USUARIOS DEL SISTEMA =====
    ws_usuarios = wb.create_sheet(title="7. Usuarios del Sistema")
    
    headers_usuarios = ['ID', 'Email', 'Nombre', 'Apellido', 'Rol', 'Total de Comentarios']
    ws_usuarios.append(headers_usuarios)
    aplicar_estilos_header(ws_usuarios, headers_usuarios)
    
    for usuario in Usuario.objects.all():
        total_comentarios = Comentario.objects.filter(usuario=usuario).count()
        ws_usuarios.append([
            usuario.id,
            usuario.email,
            usuario.nombre or 'N/A',
            usuario.apellido or 'N/A',
            usuario.get_rol_display(),
            total_comentarios
        ])
    
    aplicar_bordes(ws_usuarios, ws_usuarios.max_row, len(headers_usuarios))

    # Generar archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output